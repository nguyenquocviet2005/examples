#!/usr/bin/env python3
"""
Demonstration: Create a sample XLSX file using the document-skills
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create a new workbook
wb = Workbook()

# Get the active sheet and rename it
ws = wb.active
ws.title = "Sales Report"

# Add title
ws['A1'] = "Q4 2025 Sales Report"
ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws.merge_cells('A1:E1')
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Add headers
headers = ["Product", "Q1", "Q2", "Q3", "Total"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Add sample data
data = [
    ["Laptop", 15000, 18000, 22000],
    ["Desktop", 8000, 9500, 11000],
    ["Tablet", 5000, 6200, 7500],
    ["Phone", 12000, 14000, 16500],
]

for row_num, row_data in enumerate(data, 4):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        if col_num == 1:
            cell.font = Font(bold=True)
        else:
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal="right")

    # Add formula for total (column E)
    total_cell = ws.cell(row=row_num, column=5)
    total_cell.value = f"=SUM(B{row_num}:D{row_num})"
    total_cell.number_format = '$#,##0'
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.font = Font(bold=True)

# Add totals row
total_row = len(data) + 4
ws.cell(row=total_row, column=1).value = "TOTAL"
ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")
ws.cell(row=total_row, column=1).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

for col in range(2, 6):
    cell = ws.cell(row=total_row, column=col)
    cell.value = f"=SUM({chr(64+col)}4:{chr(64+col)}{total_row-1})"
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.number_format = '$#,##0'
    cell.alignment = Alignment(horizontal="right")

# Set column widths
ws.column_dimensions['A'].width = 15
for col in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[col].width = 12

# Add borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border

# Create a second sheet with summary statistics
summary_ws = wb.create_sheet("Summary")
summary_ws['A1'] = "Summary Statistics"
summary_ws['A1'].font = Font(size=12, bold=True)

summary_ws['A3'] = "Metric"
summary_ws['B3'] = "Value"
for cell in [summary_ws['A3'], summary_ws['B3']]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

summary_ws['A4'] = "Total Revenue"
summary_ws['B4'] = f"='Sales Report'!E{total_row}"
summary_ws['B4'].number_format = '$#,##0'

summary_ws['A5'] = "Average per Product"
summary_ws['B5'] = f"=B4/4"
summary_ws['B5'].number_format = '$#,##0'

summary_ws['A6'] = "Highest Quarter"
summary_ws['B6'] = "=MAX('Sales Report'!B4:D7)"
summary_ws['B6'].number_format = '$#,##0'

summary_ws.column_dimensions['A'].width = 20
summary_ws.column_dimensions['B'].width = 15

# Save the workbook
output_path = "/home/viet2005/workspace/fsoft/claude_skills/sample_sales_report.xlsx"
wb.save(output_path)

print(f"✅ Successfully created XLSX file: {output_path}")
print(f"\nFile contains:")
print("  • Sheet 1: 'Sales Report' - with product sales data and formulas")
print("  • Sheet 2: 'Summary' - with summary statistics using cross-sheet formulas")
print(f"\nFeatures demonstrated:")
print("  ✓ Cell formatting (bold, colors, backgrounds)")
print("  ✓ Excel formulas (SUM, cross-sheet references)")
print("  ✓ Number formatting (currency)")
print("  ✓ Multiple worksheets")
print("  ✓ Borders and alignment")
